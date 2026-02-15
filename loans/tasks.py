from celery import shared_task
import openpyxl
from .models import Customer, Loan
from .services import calculate_emi

@shared_task
def ingest_data():
    if Customer.objects.exists():
        return
    cwb = openpyxl.load_workbook('data/customer_data.xlsx')
    lwb = openpyxl.load_workbook('data/loan_data.xlsx')

    for r in cwb.active.iter_rows(min_row=2, values_only=True):
        try:
            # Skip empty rows
            if not r[0]:
                continue
            
            Customer.objects.create(
                id=r[0], first_name=r[1], last_name=r[2],
                phone_number=r[3], monthly_income=r[4],
                approved_limit=r[5], current_debt=r[6], age=30
            )
        except Exception as e:
            print(f"Error ingesting customer row {r}: {e}")

    for r in lwb.active.iter_rows(min_row=2, values_only=True):
        try:
            if not r[0]:
                continue
                
            cust_id = r[0]
            try:
                cust = Customer.objects.get(id=cust_id)
            except Customer.DoesNotExist:
                print(f"Customer {cust_id} not found for loan {r[1]}")
                continue

            emi = calculate_emi(r[2], r[3], r[4])
            Loan.objects.create(
                id=r[1], 
                customer=cust, 
                loan_amount=r[2],
                tenure=r[3], 
                interest_rate=r[4],
                monthly_installment=emi, 
                emis_paid_on_time=r[6],
                start_date=r[7],
                end_date=r[8]
            )
        except Exception as e:
            print(f"Error ingesting loan row {r}: {e}")
